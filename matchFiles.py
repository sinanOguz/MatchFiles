
from stat import S_ISREG, ST_CTIME, ST_MODE
import os, sys
import argparse
from openpyxl import Workbook
import openpyxl

def create_arg_parser():
    parser = argparse.ArgumentParser(description='match the files')
    parser.add_argument('experimentDirectory',
                    help='Path to the experiments directory.')
    parser.add_argument('videoDirectory',
                    help='Path to the videos directory.')
    parser.add_argument('spreadsheetName',
                    help='Name of the spreadsheet')
    parser.add_argument('experimentName',
                    help=' Name of the related set of experiment (e.g., experimentFormation). Creates a new page in the spreadsheet.')
    return parser


if __name__ == "__main__":
    arg_parser = create_arg_parser()
    parsed_args = arg_parser.parse_args(sys.argv[1:])
    expr_dir = parsed_args.experimentDirectory
    vid_dir = parsed_args.videoDirectory
    spreadsheet_name = parsed_args.spreadsheetName
    experiment_name = parsed_args.experimentName

if os.path.isfile(spreadsheet_name + '.xlsx'):
    print ('The spreadsheet has already been created. Overwriting the spreadsheet.')
    wb = openpyxl.load_workbook(spreadsheet_name + '.xlsx')
    ws = wb.create_sheet(experiment_name)
else:
    print ('A new spreadsheed is created.')
    wb = Workbook()
    ws = wb.active
    ws.title = experiment_name


data_expr = (os.path.join(expr_dir, fn) for fn in os.listdir(expr_dir))
data_expr = ((os.stat(path_expr), path_expr) for path_expr in data_expr)
data_expr = ((stat[ST_CTIME], path_expr)
          for stat, path_expr in data_expr if S_ISREG(stat[ST_MODE]))

data_vid = (os.path.join(vid_dir, fn) for fn in os.listdir(vid_dir))
data_vid = ((os.stat(path_vid), path_vid) for path_vid in data_vid)
data_vid = ((stat[ST_CTIME], path_vid)
          for stat, path_vid in data_vid if S_ISREG(stat[ST_MODE]))

i = 1
for cdate, path_vid in sorted(data_vid):
     ws.cell(row=i, column=1).value = os.path.basename(path_vid)
     i = i + 1

i = 1
for cdate, path_expr in sorted(data_expr):
     ws.cell(row=i, column=2).value = os.path.basename(path_expr)
     i = i + 1


wb.save(spreadsheet_name + '.xlsx')

